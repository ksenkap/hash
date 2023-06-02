import openpyxl
import random as r
import hashlib

f = open('dehash.txt')
N = [89869497194, 89037081036, 89636996427, 89669937115, 89866288994]
D={}
numbers = []
for line in f:
    n = line.split(":")
    numbers.append(int(n[1]))
    for i in range(5):
        s = int(n[1])-N[i]
        if s in D:
            D[s] += 1
        else:
            D[s] = 1
f.close()
for i in D:
    if D[i]==5:
        s=i
        print("Соль:", i)
        break

wb = openpyxl.load_workbook(filename='scoring_data_v.1.12.xlsx')
sheet = wb['A2']
sheet.cell(row=1, column=2).value = "Чистый номер"
sheet.cell(row=1, column=4).value = "Соль"
sheet.cell(row=2, column=4).value = s
clean_numbers = []
for i in range(2, len(numbers)):
    sheet.cell(row=i, column=2).value = numbers[i-2] - s
    clean_numbers.append(numbers[i-2] - s)
wb.save('scoring_data_v.1.12.xlsx')
wb.close()

f_digit_sault = open('digit.txt', 'w')
f_letter_sault = open('letter.txt', 'w')
al = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
for i in clean_numbers:
    m = i + 10234586978
    f_digit_sault.write(str(m) + '\n')
    m = str(i)
    for j in range (3):
        m += al[r.randint(0, 51)]
    f_letter_sault.write(m + '\n')
f_digit_sault.close()
f_letter_sault.close()
f_digit_sault=open("digit.txt")
f_letter_sault = open("letter.txt")
f_clean = open("clean.txt")
md5_d = open ("md5d.txt", "w")
md5_l = open ("md5l.txt", "w")
md5_c = open ("md5c.txt", "w")
sha1 = open ("sha1.txt", "w")
sha384 = open ("sha384.txt", "w")
for line in f_digit_sault:
    md5_d.write(str(hashlib.md5(line.encode('utf-8')).hexdigest()) + '\n')
for line in f_letter_sault:
    md5_l.write(str(hashlib.md5(line.encode('utf-8')).hexdigest()) + '\n')
for line in f_clean:
    md5_c.write(str(hashlib.md5(line.encode('utf-8')).hexdigest()) + '\n')
    sha1.write(str(hashlib.md5(line.encode('utf-8')).hexdigest()) + '\n')
    sha384.write(str(hashlib.md5(line.encode('utf-8')).hexdigest()) + '\n')
f_digit_sault.close()
f_letter_sault.close()
f_clean.close()
md5_c.close()
md5_d.close()
md5_l.close()
sha1.close()
sha384.close()
